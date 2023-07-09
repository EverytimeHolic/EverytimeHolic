import openpyxl
import os
import random
import csv

from bs4 import BeautifulSoup
import matplotlib.pyplot as plt

# csv 파일로 변환할것

lectures = [] # [["cur", "major", "id", "name", "num", "grade", "credit", "pf", "time", "eng", "score", "hw", "team", "mark", "att", "test"]]
score_base = "0" # 평점이 0일때 조정할 기본값(평가가 없는 강의)

# id: 과목번호
# name: 교과목명
# num: 분반
# cur: 전공이수구분
# major: 개설전공
# grade: 이수학년  (0 인경우와 빈경우가 있음)
# credit: 학점 (시간이 0인경우존재 + time도 비어있음)
# pf: 담당교수 (비어있는 경우 다수)
# time: 요일 및 교시 - [요일(월~토), 시작시간(1~14), 종료시간(1~14)] = 리스트로 구성, 분할된 경우 2개의 리스트로 구성 (비어있는경우 다수)
# eng: 외국어 강의 여부 (비어있는 경우 다수)
# score: 에타 강의평점 (0점일경우 4점으로 보정 - 미평가 강의)

# 랜덤요소
# hw: 과제 1: 없음 0: 보통 -1: 많음
# team: 팀플 1: 없음 0: 보통 -1: 많음
# mark: 채점기준 1: 너그러움 0: 보통 -1: 깐깐함
# att: 출결 반영x : 2 전자출결 : 1 복합적 : 0 지정좌석 : -1 직접호명 : -2
# test: 시험횟수 없음 : 2 한번 : 1 두번 : 0 세번 : -1 네 번이상: -2


##########################
### Trinity_EXEL_FILE ####
##########################

os.chdir(r"") # data 폴더 경로 입력
wb = openpyxl.load_workbook("Trinity_2023_2.xlsx", data_only=True)
sheet = wb['개설과목리스트']

for row in range(3, 1286): # 비고항목 제외
    lecture = []
    for column in range(1, 12):
        if column == 8 : continue # 수강인원항목 제외
        
        # 시간/학점 -> 학점
        if column == 7:
            lecture.append(sheet.cell(row=row, column=column).value[2])
        # 요일및교시 정형화
        elif column == 10:
            str = sheet.cell(row=row, column=column).value
            if str == '': # 시간표가 정해지지 않은 경우
                lecture.append(str)
            else:
                str = str.split(', ')
                time_set = []
                for item in str:
                    time = []
                    time.append(item[0]) # 요일
                    if '~' in item:
                        time.append(item[1:item.find('~')]) # 시작시간
                        time.append(item[item.find('~')+1:item.find('(')]) # 종료시간
                    else: #만약 1시간짜리라면 시작시간 = 종료시간
                        time.append(item[1:item.find('(')])
                        time.append(item[1:item.find('(')])
                    time_set.append(time)
                lecture.append(time_set)
        else:
            lecture.append(sheet.cell(row=row, column=column).value)
    lecture.append(score_base) # score자리추가
    lectures.append(lecture)

############################
### EVERYTIME_HTML_FILE ####
############################

page = open('everytime_2023_2.html', 'rt', encoding='utf-8').read() # HTML read to String
soup = BeautifulSoup(page, 'html.parser') # BS 객체 생성

scores = [] #[id, num, score]

for tr in soup.select('tr'):
    temp = list(tr)[0].get_text()
    if (temp == "" or 48 > ord(temp[1]) or ord(temp[1]) > 57): continue # 잘못된 입력값 예외처리
    id = temp[:temp.find('-')]
    
    num = temp[temp.find('-')+1:]
    
    
    star = tr.select_one('a')['title']
    if star == '0':
        star = score_base
    scores.append([id, num, star])

for score in scores:
    for lecture in lectures:
        if lecture[2] == score[0] and lecture[4] == score[1]:
            lecture[10] = score[2]


# lectures_score = [float(lecture[10]) for lecture in lectures]

# plt.scatter([i for i in range(0, len(lectures_score))], lectures_score,)
# plt.xlabel('lecture_id')
# plt.ylabel('lecture_rate')
# plt.show()

####################
### RANDOM_VALUE ###
####################

for lecture in lectures:
    if lecture[10] == score_base: # 강의평이 없는 경우
        for i in range(5):
            lecture.append("")
        
    else:
        lecture.append(random.randrange(-1, 2))
        lecture.append(random.randrange(-1, 2))
        lecture.append(random.randrange(-1, 2))
        lecture.append(random.randrange(-2, 3))
        lecture.append(random.randrange(-2, 3))
        
##################
### CSV_WRITER ###
##################

f = open('data.csv', 'w', newline='')
wr = csv.writer(f)
wr.writerow(["cur", "major", "id", "name", "num", "grade", "credit", "pf", "time", "eng", "score", "hw", "team", "mark", "att", "test"])
wr.writerows(lectures)

f.close()